import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

df = pd.read_excel("raw_results.xlsx")
df = df.dropna(subset="Name")
df = df.drop_duplicates()

#Resolving invalid scores
invalid_rows = df[
    (df["Score"]<0)|
    (df["Score"]>100)|
    (df["Score"].isna())
]

#Grade colour function
def grade_color(grade):
    if grade == "A":
        return PatternFill(start_color="90FF90", end_color="90FF90", fill_type="solid")
    elif grade == "B":
        return PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    elif grade == "C":
        return PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    elif grade == "D":
        return PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    else:
        return PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

#GPA colour Function
def gpa_color(gpa):
    if gpa >= 4.5:
        return PatternFill(start_color="90FF90", end_color="90FF90", fill_type="solid")
    elif 3.49 < gpa< 4.50:
        return PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    elif 2.39 < gpa< 3.50:
        return PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    elif 1.49 < gpa< 2.40:
        return PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    else:
        return PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
#Getting the weighted grade points
gp_list = []
remark_list = []
for score in df["Score"]:
    if score >= 70:
        gp = 5
        remark = "A"
    elif 59< score < 70:
        gp = 4
        remark = "B"
    elif 49< score <60:
        gp = 3
        remark = "C"
    elif 44< score <50:
        gp = 2
        remark = "D"
    elif 39< score <45:
        gp = 1
        remark = "E"
    else:
        gp = 0
        remark = "F"
    gp_list.append(gp)
    remark_list.append(remark)

#Adding Grade and Grade Point column
df["Grade"] = remark_list
df["GP"] = gp_list

# unit_list = df["Units"].tolist()
df["WGP"] = df["GP"] * df["Units"]

df["Name"] = df["Name"].str.strip()
df["Matric No"] = df["Matric No"].str.strip()
groups = df.groupby("Matric No")

matric_block = "nil"
gpa_dict = {}

with pd.ExcelWriter("transcripts.xlsx", engine="openpyxl") as writer:
    for matric, group in groups:
        if matric_block == matric:
            continue
        matric_block = matric
        student = groups.get_group(matric)
        wgp_sum = student["WGP"].sum()
        unit_sum = student["Units"].sum()
        gpa = round(wgp_sum/unit_sum, 2)
        # print(student)
        # print(f"Cumulative Weighted Grade Point = {wgp_sum}")
        # print(f"Total Units Taken = {unit_sum}")
        # print(f"GPA = {gpa}")
        gpa_dict[matric] = float(gpa)
        s_matric = str(matric).replace("/", "_")[:31]
        group.to_excel(writer, sheet_name=s_matric[:31], index=False)
        ws = writer.sheets[s_matric[:31]]

        #Cumulative GP
        last_row = ws.max_row + 1
        ws.merge_cells(start_row=last_row, start_column=3, end_row=last_row, end_column=8)
        cell = ws.cell(row=last_row, column=3)
        cell.value = f"Cumulative WGP = {wgp_sum}"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

        #Total Units
        last_row = ws.max_row + 1
        ws.merge_cells(start_row=last_row, start_column=3, end_row=last_row, end_column=8)
        cell = ws.cell(row=last_row, column=3)
        cell.value = f"Total Units Taken = {unit_sum}"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

        #GPA
        last_row = ws.max_row + 1
        ws.merge_cells(start_row=last_row, start_column=3, end_row=last_row, end_column=3)
        cell = ws.cell(row=last_row, column=3)
        cell.value = f"GPA = {gpa}"
        cell.font = Font(bold=True)
        cell.fill = gpa_color(gpa)
        cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length,len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
        
        grade_col = None
        header = list(ws.iter_rows(min_row = 1, max_row = 1, values_only = True))[0]
        for idx, col_name in enumerate(header, start = 1):
            if col_name == "Grade":
                grade_col = idx
                break
        for row in ws.iter_rows(min_row = 2, max_row = ws.max_row, min_col = grade_col, max_col = grade_col):
            for cell in row:
                if cell.value is not None:
                    cell.fill = grade_color(cell.value)

# Ranking according to cgpa
# df.columns = df.columns.str.strip()
# sorted_dict = dict(sorted(gpa_dict.items(),key=lambda x:x[1], reverse=True))
# subset = df[["Matric No", "Name"]]
# print(subset)
# for key, value in gpa_dict:
# print(df["Name"].dtype)
# for col in df.columns:
    # print(repr(col))
df["Course"] = df["Course"].str.strip()
group_course = df.groupby("Course")

course_score_list = []
course_score_dict = {}
for course in df["Course"]:
    course_score = group_course.get_group(course)
    course_score_list = course_score["Score"].to_list()
    course_score_dict[course] = course_score_list

print("Note the pass percentage of each course below:")
for course in course_score_dict:
    values = course_score_dict.get(course)
    len_values = len(values)
    above_50 = [value for value in values if value > 49]
    len_above_50 = len(above_50)
    print(f"{course} {round((len_above_50/len_values) * 100, 2)}")

print("\nResults have been processed and saved as transcripts.xlsx")