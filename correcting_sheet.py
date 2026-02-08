import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row+1):
        wrong_cell = sheet.cell(row, 3)
        corrected_price = wrong_cell.value * 0.9
        right_cell = sheet.cell(row, 4)
        right_cell.value = corrected_price

    ref = Reference(sheet, 
                    min_row=2, 
                    max_row=sheet.max_row, 
                    min_col=4, 
                    max_col=4)
    chart = BarChart()
    chart.add_data(ref)
    sheet.add_chart(chart, "e2")
    wb.save(filename)
filename = input("Workbook name? ")
try:
    process_workbook(filename)
except:
    print("file not found in MY BACKEND directory")